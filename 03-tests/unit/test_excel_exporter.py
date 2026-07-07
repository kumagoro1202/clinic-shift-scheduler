"""P7-6 Excel出力（exporters/excel_exporter.py）のテスト。

`external/output-design.md` 4章・5章の仕様（月間勤務表グリッド・勤務記号・
印刷レイアウト・休診日網掛け）を検証する。スケジュール・設定はテスト用の
最小データを使用し、実在データは使用しない。
"""

from __future__ import annotations

import dataclasses
import io
from pathlib import Path

import pytest
from openpyxl import load_workbook

from exporters.excel_exporter import build_grid, resolve_symbol, write_excel
from scheduler import calendar, config_loader

REPO_ROOT = Path(__file__).resolve().parents[2]
SAMPLE_CONFIG_PATH = REPO_ROOT / "config" / "samples" / "sample_clinic.yaml"
SAMPLE_STAFF_PATH = REPO_ROOT / "config" / "samples" / "sample_staff.yaml"
SAMPLE_SCHEDULE_PATH = REPO_ROOT / "config" / "samples" / "sample_schedule_202608.yaml"


@pytest.fixture(scope="module")
def config():
    return config_loader.load_config(SAMPLE_CONFIG_PATH, staff_path=SAMPLE_STAFF_PATH)


@pytest.fixture(scope="module")
def monthly(config):
    return calendar.load_monthly_schedule(SAMPLE_SCHEDULE_PATH, config)


@pytest.fixture(scope="module")
def calendar_days_all(config, monthly):
    return calendar.expand_month_all(config, monthly)


_SCHEDULE = {
    "2026-08-03": {
        "スタッフA": {
            "pattern": "early",
            "work": [
                {"area": "reception", "start": "08:30", "end": "12:30"},
                {"area": "reception", "start": "13:30", "end": "17:30"},
            ],
            "break": {"start": "12:30", "end": "13:30"},
        },
        "スタッフC": {
            "pattern": "half",
            "work": [{"area": "rehab", "start": "08:30", "end": "13:30"}],
            "break": None,
        },
    },
}


def test_resolve_symbol_uses_pattern_symbol_when_working(config, calendar_days_all):
    day = next(d for d in calendar_days_all if d.date == "2026-08-03")
    symbol = resolve_symbol(config, day, _SCHEDULE, {}, "スタッフA")
    assert symbol == "早"  # pattern=early のデフォルト記号


def test_resolve_symbol_closed_day_overrides_everything(config, calendar_days_all):
    closed_day = next(d for d in calendar_days_all if d.date == "2026-08-02")  # 日曜
    assert closed_day.day_type == "closed"
    symbol = resolve_symbol(config, closed_day, _SCHEDULE, {}, "スタッフA")
    assert symbol == "／"


def test_resolve_symbol_unpaid_full_vacation(config, calendar_days_all, monthly):
    # サンプル月次入力にはスタッフAの有給終日休暇(2026-08-05)がある
    day = next(d for d in calendar_days_all if d.date == "2026-08-05")
    vacation_lookup = {(v.staff, v.date): v for v in monthly.vacations}
    symbol = resolve_symbol(config, day, {}, vacation_lookup, "スタッフA")
    assert symbol == "有"  # paid=true が kind=full より優先


def test_resolve_symbol_unpaid_am_vacation(config, calendar_days_all, monthly):
    day = next(d for d in calendar_days_all if d.date == "2026-08-20")
    vacation_lookup = {(v.staff, v.date): v for v in monthly.vacations}
    symbol = resolve_symbol(config, day, {}, vacation_lookup, "スタッフB")
    assert symbol == "前"  # kind=am, paid=false


def test_resolve_symbol_no_assignment_no_vacation_open_day(config, calendar_days_all):
    day = next(d for d in calendar_days_all if d.date == "2026-08-03")
    symbol = resolve_symbol(config, day, {}, {}, "スタッフE")
    assert symbol == "／"


def test_build_grid_preserves_staff_order_and_counts_workdays(config, calendar_days_all):
    rows = build_grid(config, calendar_days_all, (), _SCHEDULE)
    assert [r["staff"] for r in rows] == [s.name for s in config.staff]

    row_a = next(r for r in rows if r["staff"] == "スタッフA")
    assert row_a["symbols"]["2026-08-03"] == "早"
    assert row_a["workdays"] == 1

    row_b = next(r for r in rows if r["staff"] == "スタッフB")
    assert row_b["workdays"] == 0


def test_write_excel_sheet1_layout_and_symbols(config, calendar_days_all, monthly):
    buffer = io.BytesIO()
    write_excel(
        config,
        calendar_days_all,
        monthly.vacations,
        _SCHEDULE,
        monthly.target_month,
        buffer,
    )
    buffer.seek(0)
    wb = load_workbook(buffer)

    assert wb.sheetnames == ["月間勤務表", "日別配置表"]
    ws = wb["月間勤務表"]

    assert ws.cell(row=1, column=1).value == "勤務表 2026-08"

    # 列2 = 8/1, 列3 = 8/2 ... 対応する日付列を特定する
    date_row = 2
    col_for_aug3 = next(
        col for col in range(2, ws.max_column + 1) if ws.cell(row=date_row, column=col).value == 3
    )
    weekday_row = 3
    assert ws.cell(row=weekday_row, column=col_for_aug3).value == "月"

    staff_row = next(
        r
        for r in range(weekday_row + 1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == "スタッフA"
    )
    assert ws.cell(row=staff_row, column=col_for_aug3).value == "早"

    # 出勤日数の集計列（最終列）
    total_col = ws.max_column
    assert ws.cell(row=weekday_row, column=total_col).value == "出勤日数"
    assert ws.cell(row=staff_row, column=total_col).value == 1

    # 印刷設定: A4横・1ページに収める
    assert ws.page_setup.orientation == "landscape"
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 1
    assert ws.sheet_properties.pageSetUpPr.fitToPage is True


def test_write_excel_closed_day_is_shaded(config, calendar_days_all, monthly):
    buffer = io.BytesIO()
    write_excel(
        config, calendar_days_all, monthly.vacations, _SCHEDULE, monthly.target_month, buffer
    )
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb["月間勤務表"]

    date_row, weekday_row = 2, 3
    col_for_aug2 = next(
        col for col in range(2, ws.max_column + 1) if ws.cell(row=date_row, column=col).value == 2
    )
    assert ws.cell(row=weekday_row, column=col_for_aug2).value == "日"
    fill = ws.cell(row=weekday_row, column=col_for_aug2).fill
    assert fill.start_color.rgb == "00D9D9D9"


def test_write_excel_sheet2_contains_daily_assignments(config, calendar_days_all, monthly):
    buffer = io.BytesIO()
    write_excel(
        config, calendar_days_all, monthly.vacations, _SCHEDULE, monthly.target_month, buffer
    )
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb["日別配置表"]

    assert [ws.cell(row=1, column=c).value for c in range(1, 8)] == [
        "日付",
        "曜日",
        "スタッフ",
        "パターン",
        "エリア",
        "開始",
        "終了",
    ]
    data_rows = [
        [ws.cell(row=r, column=c).value for c in range(1, 8)]
        for r in range(2, ws.max_row + 1)
    ]
    assert ["2026-08-03", "mon", "スタッフA", "early", "reception", "08:30", "12:30"] in data_rows


def test_write_excel_paper_size_and_orientation_configurable(config, calendar_days_all, monthly):
    buffer = io.BytesIO()
    write_excel(
        config,
        calendar_days_all,
        monthly.vacations,
        _SCHEDULE,
        monthly.target_month,
        buffer,
        orientation="portrait",
        paper_size="A3",
    )
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb["月間勤務表"]
    assert ws.page_setup.orientation == "portrait"
    assert int(ws.page_setup.paperSize) == 8  # A3


def test_write_excel_rejects_invalid_orientation(config, calendar_days_all, monthly):
    with pytest.raises(ValueError, match="orientation"):
        write_excel(
            config,
            calendar_days_all,
            monthly.vacations,
            _SCHEDULE,
            monthly.target_month,
            io.BytesIO(),
            orientation="sideways",
        )


def test_write_excel_rejects_invalid_paper_size(config, calendar_days_all, monthly):
    with pytest.raises(ValueError, match="paper_size"):
        write_excel(
            config,
            calendar_days_all,
            monthly.vacations,
            _SCHEDULE,
            monthly.target_month,
            io.BytesIO(),
            paper_size="B4",
        )


def test_custom_symbol_mapping_via_config(monthly):
    """Q-07回答後の設定更新のみで記号が反映されること（設定化の検証）。"""
    custom_config = config_loader.load_config(SAMPLE_CONFIG_PATH, staff_path=SAMPLE_STAFF_PATH)
    custom_symbols = config_loader.OutputSymbols(
        patterns={"early": "E", "late": "L", "half": "H"},
        vacation_full="F",
        vacation_am="AM",
        vacation_pm="PM",
        vacation_paid="P",
        none="-",
    )
    custom_config = dataclasses.replace(custom_config, output_symbols=custom_symbols)
    calendar_days_all = calendar.expand_month_all(custom_config, monthly)

    rows = build_grid(custom_config, calendar_days_all, monthly.vacations, _SCHEDULE)
    row_a = next(r for r in rows if r["staff"] == "スタッフA")
    assert row_a["symbols"]["2026-08-03"] == "E"
    assert row_a["symbols"]["2026-08-05"] == "P"  # 有給休暇


# SC-003（週40時間チェック）警告注記のテスト用スケジュール。
# スタッフAが同一ISO週（2026-08-03月〜08-06木）に29時間勤務する
# （limit_hours を下回る値に設定すれば超過、上回る値では超過しない）。
_SC003_SCHEDULE = {
    "2026-08-03": {
        "スタッフA": {
            "pattern": "early",
            "work": [
                {"area": "reception", "start": "08:30", "end": "12:30"},
                {"area": "reception", "start": "13:30", "end": "17:30"},
            ],
            "break": {"start": "12:30", "end": "13:30"},
        },
    },
    "2026-08-04": {
        "スタッフA": {
            "pattern": "early",
            "work": [
                {"area": "reception", "start": "08:30", "end": "12:30"},
                {"area": "reception", "start": "13:30", "end": "17:30"},
            ],
            "break": {"start": "12:30", "end": "13:30"},
        },
    },
    "2026-08-05": {
        "スタッフA": {
            "pattern": "early",
            "work": [
                {"area": "reception", "start": "08:30", "end": "12:30"},
                {"area": "reception", "start": "13:30", "end": "17:30"},
            ],
            "break": {"start": "12:30", "end": "13:30"},
        },
    },
    "2026-08-06": {
        "スタッフA": {
            "pattern": "half",
            "work": [{"area": "reception", "start": "08:30", "end": "13:30"}],
            "break": None,
        },
    },
}


def test_write_excel_sc003_warning_shaded_when_enabled_and_exceeded(
    config, calendar_days_all, monthly
):
    over_limit_config = dataclasses.replace(
        config,
        weekly_hours_check=config_loader.WeeklyHoursCheck(enabled=True, limit_hours=20),
    )
    buffer = io.BytesIO()
    write_excel(
        over_limit_config,
        calendar_days_all,
        monthly.vacations,
        _SC003_SCHEDULE,
        monthly.target_month,
        buffer,
    )
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb["月間勤務表"]

    date_row = 2
    staff_row = next(
        r
        for r in range(4, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == "スタッフA"
    )
    for day in (3, 4, 5, 6):
        col = next(
            c for c in range(2, ws.max_column + 1) if ws.cell(row=date_row, column=c).value == day
        )
        fill = ws.cell(row=staff_row, column=col).fill
        assert fill.start_color.rgb == "00FFC7CE", f"2026-08-{day:02d} should be shaded"

    # 超過週に属さない別スタッフの行は警告色にならない
    other_row = next(
        r
        for r in range(4, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == "スタッフB"
    )
    col_for_aug3 = next(
        c for c in range(2, ws.max_column + 1) if ws.cell(row=date_row, column=c).value == 3
    )
    assert ws.cell(row=other_row, column=col_for_aug3).fill.start_color.rgb != "00FFC7CE"


def test_write_excel_sc003_no_warning_when_disabled_even_if_exceeded(
    config, calendar_days_all, monthly
):
    # `config` fixture は weekly_hours_check.enabled=False（サンプル設定の初期値）。
    # 29時間勤務(スタッフAの限度超過と同等のデータ)でも無効時は従来通り無変化。
    buffer = io.BytesIO()
    write_excel(
        config,
        calendar_days_all,
        monthly.vacations,
        _SC003_SCHEDULE,
        monthly.target_month,
        buffer,
    )
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb["月間勤務表"]

    date_row = 2
    staff_row = next(
        r
        for r in range(4, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == "スタッフA"
    )
    for day in (3, 4, 5, 6):
        col = next(
            c for c in range(2, ws.max_column + 1) if ws.cell(row=date_row, column=c).value == day
        )
        fill = ws.cell(row=staff_row, column=col).fill
        assert fill.start_color.rgb != "00FFC7CE"
