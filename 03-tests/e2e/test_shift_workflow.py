"""P7-8 E2E: シフト生成→結果表示→手動編集→Excel出力の一連の流れ。

管理者UI（Streamlit, `02-src/ui/app.py`）を実プロセスとして起動し、
Playwright で実ブラウザ操作を行う（ARCHITECTURE.md 4章のテスト戦略・
6章 playwright(dev) の導入方針）。データソースは `config/samples/` を
一時ディレクトリへコピーし、環境変数（`CLINIC_CONFIG_PATH` 等。
`ui/pages/*.py` 参照）で差し替える。

シフト編集画面のグリッド（`st.data_editor`。Canvas 描画の
glide-data-grid 製）はセル単位の Playwright 操作が不安定要因となるため、
本テストでは画面遷移・保存アクションの往復を検証対象とし、セル単位の
編集ロジック自体は P7-5 の AppTest ユニットテスト
（`test_ui_shift_editing.py`）で担保する。

生成処理（CP-SAT、時間上限55秒。`shift_generation.DEFAULT_TIME_LIMIT_SECONDS`）
を含むため実行に数十秒かかる。flaky detection: pytest-rerunfailures により
CI 環境のタイミング起因の不安定性を検知・再試行する（`@pytest.mark.flaky`）。
"""

from __future__ import annotations

import os
import shutil
import socket
import subprocess
import sys
import time
from pathlib import Path

import pytest
from openpyxl import load_workbook
from playwright.sync_api import Page, expect

REPO_ROOT = Path(__file__).resolve().parents[2]
APP_PATH = REPO_ROOT / "02-src" / "ui" / "app.py"
SAMPLE_CONFIG_PATH = REPO_ROOT / "config" / "samples" / "sample_clinic.yaml"
SAMPLE_STAFF_PATH = REPO_ROOT / "config" / "samples" / "sample_staff.yaml"
SAMPLE_SCHEDULE_PATH = REPO_ROOT / "config" / "samples" / "sample_schedule_202608.yaml"

STARTUP_TIMEOUT_SECONDS = 30
# CP-SAT求解の時間上限（shift_generation.DEFAULT_TIME_LIMIT_SECONDS=55秒）に
# ブラウザ描画・CIの共有ランナー遅延分の余裕を加えた値。
GENERATION_TIMEOUT_MS = 90_000


def _free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return sock.getsockname()[1]


@pytest.fixture(scope="module")
def running_app(tmp_path_factory):
    """Streamlit管理者UIを実プロセスとして起動し、ベースURLを提供する。"""
    data_dir = tmp_path_factory.mktemp("e2e_data")
    config_path = data_dir / "clinic.yaml"
    staff_path = data_dir / "staff.yaml"
    schedule_path = data_dir / "schedule.yaml"
    shutil.copy(SAMPLE_CONFIG_PATH, config_path)
    shutil.copy(SAMPLE_STAFF_PATH, staff_path)
    shutil.copy(SAMPLE_SCHEDULE_PATH, schedule_path)

    port = _free_port()
    env = {
        **os.environ,
        "CLINIC_CONFIG_PATH": str(config_path),
        "CLINIC_STAFF_PATH": str(staff_path),
        "CLINIC_SCHEDULE_PATH": str(schedule_path),
        "PYTHONPATH": str(REPO_ROOT / "02-src"),
    }
    proc = subprocess.Popen(
        [
            sys.executable,
            "-m",
            "streamlit",
            "run",
            str(APP_PATH),
            "--server.headless=true",
            f"--server.port={port}",
            "--server.address=127.0.0.1",
            "--browser.gatherUsageStats=false",
        ],
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
    )

    base_url = f"http://127.0.0.1:{port}"
    deadline = time.monotonic() + STARTUP_TIMEOUT_SECONDS
    ready = False
    while time.monotonic() < deadline:
        if proc.poll() is not None:
            break
        try:
            with socket.create_connection(("127.0.0.1", port), timeout=1):
                ready = True
                break
        except OSError:
            time.sleep(0.5)

    if not ready:
        proc.terminate()
        output = proc.stdout.read() if proc.stdout else ""
        raise RuntimeError(f"Streamlitアプリの起動がタイムアウトしました:\n{output}")

    yield base_url

    proc.terminate()
    try:
        proc.wait(timeout=10)
    except subprocess.TimeoutExpired:
        proc.kill()


def _goto_page(page: Page, title: str) -> None:
    page.get_by_role("link", name=title).click()
    expect(page.get_by_role("heading", name=title)).to_be_visible(timeout=15_000)


@pytest.mark.flaky(reruns=2, reruns_delay=3)
def test_generate_display_edit_export_workflow(running_app, page: Page, tmp_path: Path):
    base_url = running_app
    page.goto(base_url)
    expect(page.get_by_role("heading", name="クリニック シフト作成システム")).to_be_visible(
        timeout=15_000
    )

    # --- シフト生成(P7-4) ---
    _goto_page(page, "シフト生成")
    page.get_by_role("button", name="生成実行").click()
    expect(page.get_by_text("シフト生成完了", exact=False)).to_be_visible(
        timeout=GENERATION_TIMEOUT_MS
    )

    # --- 結果表示(P7-4) ---
    expect(page.get_by_test_id("stDataFrame").first).to_be_visible(timeout=15_000)

    # --- 手動編集(P7-5) ---
    _goto_page(page, "シフト編集")
    expect(page.get_by_test_id("stDataFrame").first).to_be_visible(timeout=15_000)
    expect(page.get_by_role("heading", name="ハード制約違反チェック")).to_be_visible(timeout=15_000)
    page.get_by_role("button", name="保存").click()
    expect(page.get_by_text("編集結果を保存しました", exact=False)).to_be_visible(timeout=15_000)

    # --- Excel出力(P7-6) ---
    _goto_page(page, "Excel 出力")
    download_button = page.get_by_role("button", name="Excel ファイルをダウンロード")
    expect(download_button).to_be_visible(timeout=15_000)
    with page.expect_download() as download_info:
        download_button.click()
    download = download_info.value

    assert download.suggested_filename.startswith("shift_")
    assert download.suggested_filename.endswith(".xlsx")

    saved_path = tmp_path / download.suggested_filename
    download.save_as(saved_path)
    wb = load_workbook(saved_path)
    assert wb.sheetnames == ["月間勤務表", "日別配置表"]
