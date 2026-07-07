"""Excel 出力・印刷レイアウト（P7-6）。

`external/output-design.md` 4 章の仕様に基づき、月間勤務表（シート 1・印刷対象）と
日別配置表（シート 2・参考）の 2 シート構成で出力する。シート 1 は A4 横 1 ページに
収まることを必須条件とする。勤務記号は `Config.output_symbols`（Q-07 仮デフォルト。
`output.symbols` で上書き可能）を使用する。出力対象のスケジュールは呼び出し側
（UI 層）が渡したもの（P7-5 手動編集後の最新値であること）であり、本モジュールは
生成・編集ロジックを持たない。SC-003 週40時間チェック（`external/output-design.md`
6 章）が有効な場合、超過週に属するセルへ警告色を付ける（無効時は従来通り無変化）。
"""

from __future__ import annotations

from pathlib import Path
from typing import BinaryIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from exporters.csv_exporter import build_rows
from scheduler.calendar import CalendarDay, DateVacation
from scheduler.config_loader import Config, parse_time
from scheduler.result import _iso_week_key

WEEKDAY_LABELS = {
    "mon": "月",
    "tue": "火",
    "wed": "水",
    "thu": "木",
    "fri": "金",
    "sat": "土",
    "sun": "日",
}

_CLOSED_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_SC003_WARNING_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_CENTER = Alignment(horizontal="center")

_ORIENTATIONS = ("landscape", "portrait")
_PAPER_SIZES = {"A4": "9", "A3": "8"}  # openpyxl paperSize コード


def resolve_symbol(
    config: Config,
    day: CalendarDay,
    schedule: dict,
    vacation_lookup: dict[tuple[str, str], DateVacation],
    staff_name: str,
) -> str:
    """スタッフ 1 名・1 日分の勤務表記号を解決する（5 章の記号体系）。

    優先順位: 休診日 > 勤務あり（パターン記号） > 休暇（記号は kind 別。
    ただし paid=true は kind によらず「有」を優先） > 配置なし。
    """
    symbols = config.output_symbols
    if day.day_type == "closed":
        return symbols.none

    entry = schedule.get(day.date, {}).get(staff_name)
    if entry is not None:
        return symbols.patterns.get(entry["pattern"], entry["pattern"])

    vacation = vacation_lookup.get((staff_name, day.date))
    if vacation is not None:
        return symbols.vacation_symbol(vacation.kind, vacation.paid)

    return symbols.none


def _sc003_warning_cells(
    config: Config,
    calendar_days_all: tuple[CalendarDay, ...],
    schedule: dict,
) -> set[tuple[str, str]]:
    """SC-003週40時間超過（`output-design.md` 6章）の対象セル（スタッフ, 日付）集合。

    `config.weekly_hours_check` が無効な場合は空集合を返す（従来通り無変化）。
    `result.validate_hard_constraints` と同じ週実働時間の集計基準を用い、
    上限超過の週に属する全ての勤務日を対象とする。
    """
    if not config.weekly_hours_check.enabled:
        return set()

    limit_minutes = config.weekly_hours_check.limit_hours * 60
    weekly_minutes: dict[tuple[str, str], int] = {}
    dates_by_week: dict[tuple[str, str], list[str]] = {}
    for day in calendar_days_all:
        week_key = _iso_week_key(day.date)
        for staff_name, entry in schedule.get(day.date, {}).items():
            work_minutes = sum(
                parse_time(seg["end"]) - parse_time(seg["start"])
                for seg in entry.get("work") or []
            )
            if not work_minutes:
                continue
            key = (staff_name, week_key)
            weekly_minutes[key] = weekly_minutes.get(key, 0) + work_minutes
            dates_by_week.setdefault(key, []).append(day.date)

    cells: set[tuple[str, str]] = set()
    for key, minutes in weekly_minutes.items():
        if minutes > limit_minutes:
            staff_name = key[0]
            cells.update((staff_name, date) for date in dates_by_week[key])
    return cells


def build_grid(
    config: Config,
    calendar_days_all: tuple[CalendarDay, ...],
    vacations: tuple[DateVacation, ...],
    schedule: dict,
) -> list[dict]:
    """行 = スタッフの月間勤務表グリッドを組み立てる。

    戻り値は `[{"staff": str, "symbols": {date: 記号}, "workdays": int}, ...]`
    （`staff` は `config.staff` の順序を維持）。`workdays` は出勤日数
    （SC-004 の公平性確認用の集計列）。
    """
    vacation_lookup = {(v.staff, v.date): v for v in vacations}
    rows = []
    for staff in config.staff:
        symbols_by_date = {}
        workdays = 0
        for day in calendar_days_all:
            symbols_by_date[day.date] = resolve_symbol(
                config, day, schedule, vacation_lookup, staff.name
            )
            if staff.name in schedule.get(day.date, {}):
                workdays += 1
        rows.append({"staff": staff.name, "symbols": symbols_by_date, "workdays": workdays})
    return rows


def _write_monthly_sheet(
    ws: Worksheet,
    target_month: str,
    calendar_days_all: tuple[CalendarDay, ...],
    grid_rows: list[dict],
    orientation: str,
    paper_size: str,
    warning_cells: set[tuple[str, str]],
) -> None:
    ws.title = "月間勤務表"
    ws.cell(row=1, column=1, value=f"勤務表 {target_month}").font = _HEADER_FONT

    date_row, weekday_row = 2, 3
    ws.cell(row=date_row, column=1, value="")
    ws.cell(row=weekday_row, column=1, value="")
    for col, day in enumerate(calendar_days_all, start=2):
        day_of_month = int(day.date[-2:])
        date_cell = ws.cell(row=date_row, column=col, value=day_of_month)
        date_cell.alignment = _CENTER
        weekday_cell = ws.cell(row=weekday_row, column=col, value=WEEKDAY_LABELS[day.weekday])
        weekday_cell.alignment = _CENTER
        if day.day_type == "closed":
            date_cell.fill = _CLOSED_FILL
            weekday_cell.fill = _CLOSED_FILL

    total_col = len(calendar_days_all) + 2
    ws.cell(row=weekday_row, column=total_col, value="出勤日数").font = _HEADER_FONT

    for r, row in enumerate(grid_rows, start=weekday_row + 1):
        ws.cell(row=r, column=1, value=row["staff"])
        for col, day in enumerate(calendar_days_all, start=2):
            cell = ws.cell(row=r, column=col, value=row["symbols"][day.date])
            cell.alignment = _CENTER
            if day.day_type == "closed":
                cell.fill = _CLOSED_FILL
            elif (row["staff"], day.date) in warning_cells:
                cell.fill = _SC003_WARNING_FILL
        ws.cell(row=r, column=total_col, value=row["workdays"]).alignment = _CENTER

    ws.column_dimensions["A"].width = 12
    for col in range(2, total_col):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 4
    ws.column_dimensions[ws.cell(row=1, column=total_col).column_letter].width = 10

    # 印刷設定: A4横・1ページに収める（`external/output-design.md` 4章）
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = _PAPER_SIZES[paper_size]
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True


def _write_daily_sheet(
    ws: Worksheet,
    calendar_days: tuple[CalendarDay, ...],
    schedule: dict,
) -> None:
    ws.title = "日別配置表"
    headers = ("date", "weekday", "staff", "pattern", "area", "start", "end")
    header_labels = ("日付", "曜日", "スタッフ", "パターン", "エリア", "開始", "終了")
    for col, label in enumerate(header_labels, start=1):
        ws.cell(row=1, column=col, value=label).font = _HEADER_FONT

    for r, row in enumerate(build_rows(schedule, calendar_days), start=2):
        for col, key in enumerate(headers, start=1):
            ws.cell(row=r, column=col, value=row[key])


def write_excel(
    config: Config,
    calendar_days_all: tuple[CalendarDay, ...],
    vacations: tuple[DateVacation, ...],
    schedule: dict,
    target_month: str,
    output: str | Path | BinaryIO,
    orientation: str = "landscape",
    paper_size: str = "A4",
) -> None:
    """月間勤務表 Excel を出力する（シート1: 印刷対象・シート2: 参考）。

    `orientation` は "landscape"/"portrait"、`paper_size` は "A4"/"A3"。
    `output` はファイルパスまたはファイルライクオブジェクト（`BytesIO` 等）。
    """
    if orientation not in _ORIENTATIONS:
        raise ValueError(f"orientation は {_ORIENTATIONS} のいずれかである必要があります")
    if paper_size not in _PAPER_SIZES:
        raise ValueError(f"paper_size は {tuple(_PAPER_SIZES)} のいずれかである必要があります")

    grid_rows = build_grid(config, calendar_days_all, vacations, schedule)
    warning_cells = _sc003_warning_cells(config, calendar_days_all, schedule)

    wb = Workbook()
    monthly_sheet = wb.active
    _write_monthly_sheet(
        monthly_sheet,
        target_month,
        calendar_days_all,
        grid_rows,
        orientation,
        paper_size,
        warning_cells,
    )

    open_days = tuple(day for day in calendar_days_all if day.day_type != "closed")
    daily_sheet = wb.create_sheet()
    _write_daily_sheet(daily_sheet, open_days, schedule)

    wb.save(output)
